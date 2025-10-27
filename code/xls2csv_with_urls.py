from openpyxl import load_workbook
import csv
import sys

if len(sys.argv) != 3:
  print("Use prog xls-in csv-out")


def xlsx_to_csv_with_links(xlsx_path, csv_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=False):
            # Extract values or hyperlinks
            out_row = []
            for cell in row:
                if cell.hyperlink:
                    out_row.append(cell.hyperlink.target)
                elif cell.value is not None:
                    out_row.append(str(cell.value).strip())
                else:
                    out_row.append("")
            
            # Skip rows that are completely empty
            if any(v.strip() for v in out_row if isinstance(v, str)):
                writer.writerow(out_row)

xlsx_to_csv_with_links(sys.argv[1], sys.argv[2])
